"""
Ingest questionnaire answers into the approved answers database.

Supports:
- CSV files (question, answer, category, framework columns)
- XLSX files (SIG-style questionnaires with Q&A columns)
- JSON files (structured Q&A exports)

Usage:
    python ingest_answers.py --file questionnaire.csv --framework SIG --approved-by "John Smith"
    python ingest_answers.py --file soc2_responses.xlsx --framework SOC2
    python ingest_answers.py --dir ./questionnaires/ --framework vendor_questionnaire
"""

import argparse
import csv
import json
import os
import sys
import uuid
from datetime import datetime

import boto3

ANSWERS_TABLE = "scf-agent-approved-answers"
REGION = "us-east-1"


def load_csv(filepath: str) -> list:
    """Load Q&A pairs from a CSV file."""
    pairs = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Look for common column names
            question = (
                row.get("question") or row.get("Question") or
                row.get("question_text") or row.get("Control Question") or
                row.get("Requirement") or ""
            )
            answer = (
                row.get("answer") or row.get("Answer") or
                row.get("answer_text") or row.get("Response") or
                row.get("Implementation") or ""
            )
            category = (
                row.get("category") or row.get("Category") or
                row.get("Domain") or row.get("Section") or ""
            )

            if question and answer:
                pairs.append({
                    "question_text": question.strip(),
                    "answer_text": answer.strip(),
                    "category": category.strip().lower().replace(" ", "_") if category else "general",
                })
    return pairs


def load_xlsx(filepath: str) -> list:
    """Load Q&A pairs from an Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required for XLSX files. Install: pip install openpyxl")
        sys.exit(1)

    pairs = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Find header row
    headers = [cell.value for cell in ws[1] if cell.value]
    headers_lower = [h.lower() if h else "" for h in headers]

    # Find question and answer columns
    q_col = None
    a_col = None
    cat_col = None

    for i, h in enumerate(headers_lower):
        if any(kw in h for kw in ["question", "requirement", "control"]):
            q_col = i
        elif any(kw in h for kw in ["answer", "response", "implementation", "description"]):
            a_col = i
        elif any(kw in h for kw in ["category", "domain", "section", "group"]):
            cat_col = i

    if q_col is None or a_col is None:
        print(f"WARNING: Could not identify question/answer columns in {filepath}")
        print(f"  Headers found: {headers}")
        return pairs

    for row in ws.iter_rows(min_row=2, values_only=True):
        question = str(row[q_col] or "").strip()
        answer = str(row[a_col] or "").strip()
        category = str(row[cat_col] or "general").strip().lower().replace(" ", "_") if cat_col else "general"

        if question and answer and len(answer) > 10:
            pairs.append({
                "question_text": question,
                "answer_text": answer,
                "category": category,
            })

    return pairs


def load_json(filepath: str) -> list:
    """Load Q&A pairs from a JSON file."""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        return data
    elif "answers" in data:
        return data["answers"]
    elif "questions" in data:
        return data["questions"]
    return []


def ingest(pairs: list, framework: str, approved_by: str, source_file: str):
    """Write Q&A pairs to DynamoDB."""
    dynamodb = boto3.resource("dynamodb", region_name=REGION)
    table = dynamodb.Table(ANSWERS_TABLE)

    count = 0
    with table.batch_writer() as batch:
        for pair in pairs:
            item = {
                "answer_id": str(uuid.uuid4()),
                "question_text": pair["question_text"],
                "answer_text": pair["answer_text"],
                "category": pair.get("category", "general"),
                "source_framework": framework.upper(),
                "source_document": os.path.basename(source_file),
                "approved_by": approved_by,
                "approved_date": datetime.utcnow().strftime("%Y-%m-%d"),
                "status": "APPROVED",
                "tags": pair.get("tags", ""),
                "scf_control_ids": pair.get("scf_control_ids", ""),
            }
            # Remove empty strings
            item = {k: v for k, v in item.items() if v}
            batch.put_item(Item=item)
            count += 1

    return count


def main():
    parser = argparse.ArgumentParser(description="Ingest questionnaire answers")
    parser.add_argument("--file", help="Path to questionnaire file (CSV, XLSX, or JSON)")
    parser.add_argument("--dir", help="Directory of questionnaire files to ingest")
    parser.add_argument("--framework", required=True,
                        help="Source framework (SIG, SOC2, HITRUST, HIPAA, vendor_questionnaire, RFP)")
    parser.add_argument("--approved-by", default="auto-import",
                        help="Who approved these answers")

    args = parser.parse_args()

    files = []
    if args.file:
        files.append(args.file)
    elif args.dir:
        for f in os.listdir(args.dir):
            if f.endswith((".csv", ".xlsx", ".json")):
                files.append(os.path.join(args.dir, f))
    else:
        print("ERROR: Provide --file or --dir")
        sys.exit(1)

    total = 0
    for filepath in files:
        print(f"Processing: {filepath}")
        ext = os.path.splitext(filepath)[1].lower()

        if ext == ".csv":
            pairs = load_csv(filepath)
        elif ext == ".xlsx":
            pairs = load_xlsx(filepath)
        elif ext == ".json":
            pairs = load_json(filepath)
        else:
            print(f"  Skipping unsupported format: {ext}")
            continue

        if pairs:
            count = ingest(pairs, args.framework, args.approved_by, filepath)
            print(f"  Loaded {count} Q&A pairs")
            total += count
        else:
            print(f"  No Q&A pairs found")

    print(f"\n✅ Total: {total} answers ingested into {ANSWERS_TABLE}")


if __name__ == "__main__":
    main()
